"""
Exportador para formato Excel (.xlsx).
Requer: pip install openpyxl
"""

from typing import List, Dict
from .base_exporter import BaseExporter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelExporter(BaseExporter):
    """Exporta histórias em planilha Excel"""

    def export(self, stories: List[Dict]) -> bytes:
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Histórias"

        # Headers
        headers = [
            "ID", "Título", "Complexidade",
            "Regras de Negócio", "APIs/Serviços", "Objetivos",
            "Critérios de Aceitação", "Criado em", "Atualizado em"
        ]
        ws.append(headers)

        # Estilizar header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adicionar dados
        for story in stories:
            ws.append([
                story['id'],
                story['titulo'],
                story['complexidade'],
                "\n".join(story.get('regras_negocio', [])),
                "\n".join(story.get('apis_servicos', [])),
                "\n".join(story.get('objetivos', [])),
                "\n".join(story.get('criterios_aceitacao', [])),
                story.get('created_at', ''),
                story.get('updated_at', '')
            ])

        # Ajustar larguras
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em bytes
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def get_filename(self, base_name: str, timestamp: str) -> str:
        return f"{base_name}-{timestamp}.xlsx"

    def get_mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
